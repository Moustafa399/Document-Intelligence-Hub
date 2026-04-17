"""Excel and CSV file extractor."""
import os
import re
import json
import csv
import openpyxl
import xlrd

from helpers.file_utils import create_document_folder, save_text, save_metadata, save_tables
from helpers.table_utils import (
    preprocess_excel_data, 
    clean_numeric_values, 
    format_table_as_markdown, 
    detect_numeric_columns
)


def extract_excel_old(file_path):
    """
    Extract data from old Excel files (.xls) using xlrd.
    """
    doc_id, base, text_dir, img_dir = create_document_folder(file_path)

    workbook = xlrd.open_workbook(file_path)
    
    text = f"EXCEL WORKBOOK: {os.path.basename(file_path)}\n"
    text += f"Total Sheets: {workbook.nsheets}\n\n"
    
    tables_data = []
    images = []
    
    for sheet_idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_idx)
        sheet_name = sheet.name
        
        text += f"\n{'='*60}\n"
        text += f"SHEET {sheet_idx + 1}: {sheet_name}\n"
        text += f"{'='*60}\n\n"
        
        # Get sheet dimensions
        nrows = sheet.nrows
        ncols = sheet.ncols
        
        if nrows == 0 or ncols == 0:
            text += "[Empty Sheet]\n"
            continue
        
        # Extract table data from sheet
        table_data = []
        for row_idx in range(nrows):
            row_data = []
            for col_idx in range(ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                
                # Handle different cell types
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = ""
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    # Convert date to string
                    try:
                        date_tuple = xlrd.xldate_as_tuple(value, workbook.datemode)
                        value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                    except:
                        value = str(value)
                else:
                    value = str(value).strip()
                
                row_data.append(value)
            table_data.append(row_data)
        
        # Preprocess the table data
        table_data = preprocess_excel_data(table_data)
        table_data = [
            [clean_numeric_values(cell) for cell in row]
            for row in table_data
        ]
        
        if table_data:
            # Separate headers from data rows
            headers = table_data[0] if table_data else []
            data_rows = table_data[1:] if len(table_data) > 1 else []
            
            tables_data.append({
                "sheet": sheet_name,
                "sheet_index": sheet_idx + 1,
                "rows": len(data_rows),
                "columns": len(headers) if headers else 0,
                "headers": headers,
                "data": data_rows
            })
            
            text += f"[TABLE: {sheet_name}]\n"
            text += f"Dimensions: {len(table_data)} rows × {len(table_data[0]) if table_data else 0} columns\n\n"
            text += format_table_as_markdown(table_data)
            text += "\n\n"
            
            numeric_cols = detect_numeric_columns(table_data)
            if numeric_cols:
                text += f"Numeric columns detected: {', '.join(numeric_cols)}\n\n"
    
    if tables_data:
        save_tables(base, tables_data)
        print(f"📊 Found {len(tables_data)} sheet(s) with data in Excel (.xls)")
    
    save_text(text_dir, text)
    save_metadata(base, {
        "source": "excel",
        "file_format": ".xls (Excel 97-2003)",
        "sheets": workbook.nsheets,
        "tables_found": len(tables_data),
        "images_found": len(images)
    })
    
    return base, images, doc_id, "excel"


def extract_excel(file_path):
    """
    Extract data from Excel files (.xlsx, .xls).
    Automatically detects format and uses appropriate library.
    """
    file_ext = os.path.splitext(file_path)[1].lower()
    
    # Old .xls format - use xlrd
    if file_ext == ".xls":
        return extract_excel_old(file_path)
    
    # Modern .xlsx format - use openpyxl
    doc_id, base, text_dir, img_dir = create_document_folder(file_path)

    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    text = f"EXCEL WORKBOOK: {os.path.basename(file_path)}\n"
    text += f"Total Sheets: {len(workbook.sheetnames)}\n\n"
    
    tables_data = []
    images = []
    charts_data = []
    
    # Load workbook again without data_only to access charts
    try:
        workbook_with_charts = openpyxl.load_workbook(file_path, data_only=False)
    except:
        workbook_with_charts = None
    
    for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
        sheet = workbook[sheet_name]
        
        text += f"\n{'='*60}\n"
        text += f"SHEET {sheet_idx}: {sheet_name}\n"
        text += f"{'='*60}\n\n"
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 0 or max_col == 0:
            text += "[Empty Sheet]\n"
            continue
        
        # Extract table data from sheet
        table_data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value).strip())
            table_data.append(row_data)
        
        # Preprocess the table data
        table_data = preprocess_excel_data(table_data)
        table_data = [
            [clean_numeric_values(cell) for cell in row]
            for row in table_data
        ]
        
        if table_data:
            # Separate headers from data rows
            headers = table_data[0] if table_data else []
            data_rows = table_data[1:] if len(table_data) > 1 else []
            
            tables_data.append({
                "sheet": sheet_name,
                "sheet_index": sheet_idx,
                "rows": len(data_rows),
                "columns": len(headers) if headers else 0,
                "headers": headers,
                "data": data_rows
            })
            
            text += f"[TABLE: {sheet_name}]\n"
            text += f"Dimensions: {len(table_data)} rows × {len(table_data[0]) if table_data else 0} columns\n\n"
            text += format_table_as_markdown(table_data)
            text += "\n\n"
            
            numeric_cols = detect_numeric_columns(table_data)
            if numeric_cols:
                text += f"Numeric columns detected: {', '.join(numeric_cols)}\n\n"
        
        # Extract charts from sheet
        if workbook_with_charts:
            try:
                chart_sheet = workbook_with_charts[sheet_name]
                if hasattr(chart_sheet, '_charts') and chart_sheet._charts:
                    text += f"\n[CHARTS FOUND: {len(chart_sheet._charts)} chart(s)]\n"
                    
                    for chart_idx, chart in enumerate(chart_sheet._charts, 1):
                        chart_info = {
                            "sheet": sheet_name,
                            "sheet_index": sheet_idx,
                            "chart_index": chart_idx,
                            "chart_type": type(chart).__name__,
                            "title": "",
                            "data_series": []
                        }
                        
                        # Get chart title
                        if hasattr(chart, 'title') and chart.title:
                            if hasattr(chart.title, 'text'):
                                chart_info["title"] = chart.title.text or ""
                            elif isinstance(chart.title, str):
                                chart_info["title"] = chart.title
                        
                        # Get chart type display name
                        chart_type_map = {
                            "BarChart": "Bar Chart",
                            "BarChart3D": "3D Bar Chart",
                            "LineChart": "Line Chart",
                            "LineChart3D": "3D Line Chart",
                            "PieChart": "Pie Chart",
                            "PieChart3D": "3D Pie Chart",
                            "AreaChart": "Area Chart",
                            "AreaChart3D": "3D Area Chart",
                            "ScatterChart": "Scatter/XY Chart",
                            "DoughnutChart": "Doughnut Chart",
                            "RadarChart": "Radar Chart",
                            "BubbleChart": "Bubble Chart",
                            "StockChart": "Stock Chart",
                            "SurfaceChart": "Surface Chart",
                            "SurfaceChart3D": "3D Surface Chart"
                        }
                        chart_info["chart_type_display"] = chart_type_map.get(
                            chart_info["chart_type"], 
                            chart_info["chart_type"]
                        )
                        
                        # Get data series information
                        if hasattr(chart, 'series'):
                            for series_idx, series in enumerate(chart.series):
                                series_info = {"index": series_idx + 1}
                                if hasattr(series, 'title') and series.title:
                                    series_info["name"] = str(series.title)
                                if hasattr(series, 'val') and series.val:
                                    if hasattr(series.val, 'numRef') and series.val.numRef:
                                        series_info["data_range"] = series.val.numRef.f
                                if hasattr(series, 'cat') and series.cat:
                                    if hasattr(series.cat, 'numRef') and series.cat.numRef:
                                        series_info["category_range"] = series.cat.numRef.f
                                    elif hasattr(series.cat, 'strRef') and series.cat.strRef:
                                        series_info["category_range"] = series.cat.strRef.f
                                chart_info["data_series"].append(series_info)
                        
                        charts_data.append(chart_info)
                        
                        text += f"\n  Chart {chart_idx}: {chart_info['chart_type_display']}\n"
                        if chart_info["title"]:
                            text += f"    Title: {chart_info['title']}\n"
                        if chart_info["data_series"]:
                            text += f"    Data Series: {len(chart_info['data_series'])}\n"
                        
            except Exception as e:
                print(f"⚠️ Could not extract charts from sheet {sheet_name}: {e}")
        
        # Extract images from sheet
        if hasattr(sheet, '_images'):
            for img_idx, img in enumerate(sheet._images, 1):
                try:
                    img_path = os.path.join(img_dir, f"sheet_{sheet_idx}_img_{img_idx}.png")
                    img.image.save(img_path)
                    images.append(img_path)
                except Exception as e:
                    print(f"⚠️ Could not extract image from sheet {sheet_name}: {e}")
    
    if tables_data:
        save_tables(base, tables_data)
        print(f"📊 Found {len(tables_data)} sheet(s) with data in Excel (.xlsx)")
    
    if charts_data:
        charts_dir = os.path.join(base, "charts")
        os.makedirs(charts_dir, exist_ok=True)
        with open(os.path.join(charts_dir, "charts.json"), "w", encoding="utf-8") as f:
            json.dump(charts_data, f, indent=2, ensure_ascii=False)
        print(f"📈 Found {len(charts_data)} chart(s) in Excel workbook")
    
    save_text(text_dir, text)
    save_metadata(base, {
        "source": "excel",
        "file_format": ".xlsx",
        "sheets": len(workbook.sheetnames),
        "tables_found": len(tables_data),
        "charts_found": len(charts_data),
        "images_found": len(images)
    })
    
    return base, images, doc_id, "excel"


def extract_csv(file_path):
    """
    Extract data from CSV files.
    CSV files are treated as a single table.
    """
    doc_id, base, text_dir, img_dir = create_document_folder(file_path)
    
    text = f"CSV FILE: {os.path.basename(file_path)}\n\n"
    
    tables_data = []
    images = []
    
    # Try different encodings
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    table_data = None
    encoding_used = None
    delimiter = ','
    
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding, newline='') as f:
                # Detect delimiter
                sample = f.read(4096)
                f.seek(0)
                
                sniffer = csv.Sniffer()
                try:
                    dialect = sniffer.sniff(sample)
                    delimiter = dialect.delimiter
                except:
                    delimiter = ','
                
                reader = csv.reader(f, delimiter=delimiter)
                table_data = list(reader)
                encoding_used = encoding
                break
        except Exception:
            continue
    
    if table_data is None:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            text += f"[Error reading CSV file properly]\n{content[:1000]}"
            save_text(text_dir, text)
            save_metadata(base, {
                "source": "csv",
                "error": "Could not parse CSV",
                "tables_found": 0
            })
            return base, images, doc_id, "csv"
    
    # Remove completely empty rows
    table_data = [row for row in table_data if any(cell.strip() for cell in row if cell)]
    
    if not table_data:
        text += "[Empty CSV file]\n"
        save_text(text_dir, text)
        save_metadata(base, {
            "source": "csv",
            "encoding": encoding_used,
            "tables_found": 0
        })
        return base, images, doc_id, "csv"
    
    num_rows = len(table_data)
    num_cols = len(table_data[0]) if table_data else 0
    
    # Separate headers from data rows
    headers = table_data[0] if table_data else []
    data_rows = table_data[1:] if len(table_data) > 1 else []
    
    tables_data.append({
        "name": os.path.basename(file_path),
        "rows": len(data_rows),
        "columns": len(headers),
        "delimiter": delimiter,
        "encoding": encoding_used,
        "headers": headers,
        "data": data_rows
    })
    
    text += f"[TABLE: {os.path.basename(file_path)}]\n"
    text += f"Dimensions: {num_rows} rows × {num_cols} columns\n"
    text += f"Delimiter: '{delimiter}' | Encoding: {encoding_used}\n\n"
    text += format_table_as_markdown(table_data)
    text += "\n\n"
    
    numeric_cols = detect_numeric_columns(table_data)
    if numeric_cols:
        text += f"Numeric columns detected: {', '.join(numeric_cols)}\n\n"
    
    # Detect data characteristics
    has_header = True
    if num_rows > 1:
        first_row = table_data[0]
        numeric_in_first = sum(1 for cell in first_row if cell and re.match(r'^-?\d+\.?\d*$', cell.strip()))
        if numeric_in_first > len(first_row) * 0.5:
            has_header = False
    
    text += f"Header row detected: {has_header}\n"
    
    if tables_data:
        save_tables(base, tables_data)
        print(f"📊 CSV file parsed successfully: {num_rows} rows × {num_cols} columns")
    
    save_text(text_dir, text)
    save_metadata(base, {
        "source": "csv",
        "encoding": encoding_used,
        "delimiter": delimiter,
        "rows": num_rows,
        "columns": num_cols,
        "has_header": has_header,
        "tables_found": len(tables_data)
    })
    
    return base, images, doc_id, "csv"
